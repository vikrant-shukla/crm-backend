from io import BytesIO
import json
import random
from django.http import HttpResponse
import openpyxl
from rest_framework.response import Response
from mysite import settings
from customer.models import  Followup, Language, Otp,Representatives, Selected,UserTable, Vendor,Candidate,Jobdescription
from customer.utils import limit_off
from .serializers import FollowupSerializer, LanguageSerializer, SelectedSerializer, UserTableSerializer,OtpVerificationSerializer,JobdescriptionSerializer, VendorSerializer,CandidateSerializer, RepresentativesSerializer,AuthTokenSerializer,SetNewPasswordSerializer
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework import status, generics
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.core.mail import send_mail
from rest_framework.generics import UpdateAPIView
from openpyxl.utils import get_column_letter
from django.contrib.auth.hashers import make_password
from customer import serializers

class RegisterAPI(APIView):
    def post(self, request):
        serializer = UserTableSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'successfully registered'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors,status = status.HTTP_400_BAD_REQUEST)

class LoginAPI(TokenObtainPairView):
    permission_classes = (AllowAny,)
    serializer_class = AuthTokenSerializer
    
class SentMailView(APIView):
    """Api to sent the otp to user mail  id to reset the password"""
    def post(self, request):
        """sending the otp to user mail id"""
        try:
            mail = UserTable.objects.get(email=request.data['email'].lower())
        except:
            return Response({'error': 'Email does not exits.'}, status=status.HTTP_404_NOT_FOUND )
        if Otp.objects.filter(email=mail).exists:
            Otp.objects.filter(email=mail).delete()
        otp = Otp.objects.create(email=mail)
        otp.otp = random.randint(100000, 999999)
        otp.save()
        subject = 'Reset Your Password'
        body = f'This is your OTP to reset password {otp.otp}'
        try:
            send_mail(subject, body, settings.EMAIL_HOST_USER, [mail.email], fail_silently=False)
            return Response({"status": "mail sent "}, status=status.HTTP_201_CREATED)
        except:
            return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
        
class OtpVerification(APIView):
    serializer_class = OtpVerificationSerializer
    def post(self, request):
        serializer = self.serializer_class(data=request.data, instance=request.data)
        if serializer.is_valid(raise_exception=True):
            return Response({"otp": "verified"}, status=status.HTTP_200_OK)
        return Response({"otp": "please generate otp again"}, status=status.HTTP_400_BAD_REQUEST)
    
class ResetPasswordview(generics.UpdateAPIView):
    """Api to reset the password and storing the new password into database"""
    if OtpVerificationSerializer:
        serializer_class = SetNewPasswordSerializer
        def post(self, request, *args, **kwargs):
            """saving the new password of the user into database"""
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            request_email = request.data['email']
            user_object = UserTable.objects.get(email=request_email)
            if Otp.objects.filter(email_id=user_object.pk).exists():
                if UserTable.objects.get(email=request_email):
                    user_object.password = make_password(request.data['password'])
                    user_object.save()
                    otp_del = Otp.objects.filter(email=user_object.id)
                    otp_del.delete()
                    return Response({'status': 'password successfully changed'}, status=status.HTTP_201_CREATED)
                return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
            return Response({"status": "password successfully changed!!!"}, status=status.HTTP_400_BAD_REQUEST)
        
class VendorAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self,request):
        serializers = limit_off(Vendor,request, VendorSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        serializers = VendorSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status= 201)
        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request):
        try:
            query_parameter = Vendor.objects.get(id =request.query_params['id'] )
            data = request.data
            for key, data_value in data.items(): 
                query_parameter.__dict__[key] = data_value
            serializers = VendorSerializer(query_parameter)
            return Response(serializers.data,status= status.HTTP_201_CREATED)
        except:
            return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)

class RepresentativesAPIView(UpdateAPIView):
    # queryset= Representatives.objects.all()
    permission_classes = (IsAuthenticated,)
    def get(self,request):
        serializers = limit_off(model=Representatives,request=request, serial=RepresentativesSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        serializers = RepresentativesSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request):
        try:
            query_parameter = Representatives.objects.get(id =request.query_params['id'] )
            data = request.data
            for key, data_value in data.items(): 
                query_parameter.__dict__[key] = data_value
                    
            query_parameter.save()
            serializers = RepresentativesSerializer(query_parameter)
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        except:
            return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
                 
class LanguageAPIView(UpdateAPIView):
    
    permission_classes = (IsAuthenticated,)
    def get(self,request):
        serializers = limit_off(Language,request, LanguageSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        if Language.objects.filter(language =  request.data['language'].lower()).exists():
            return Response({"message":"language already exists"},status=  status.HTTP_400_BAD_REQUEST)
        serializers = LanguageSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
    
class CandidateAPIView(UpdateAPIView):    
    permission_classes = (IsAuthenticated,)
    
    def get(self,request):
        serializers = limit_off(Candidate,request, CandidateSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        serializers = CandidateSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request):
        try:
            query_parameter = Candidate.objects.get(id =request.query_params['id'] )
            query_parameter_l = Language.objects.get(id =request.query_params['id1'] )
            data = request.data
            for key, data_value in data.items(): 
                query_parameter.__dict__[key] = data_value
                query_parameter_l.__dict__[key] = data_value
                    
            query_parameter.save()
            query_parameter_l.save()
            serializers = CandidateSerializer(query_parameter)
            serializers = LanguageSerializer(query_parameter_l)
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        except:
            return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)
    
class JobdescriptionAPIView(UpdateAPIView):    
    permission_classes = (IsAuthenticated,)
    
    def get(self,request):
        serializers = limit_off(Jobdescription,request, JobdescriptionSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        serializers = JobdescriptionSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status= 201)
        return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)

    def patch(self, request):
        try:
            query_parameter = Jobdescription.objects.get(id =request.query_params['id'] )
            query_parameter_l = Candidate.objects.get(id =request.query_params['id1'] )
            data = request.data
            for key, data_value in data.items(): 
                query_parameter.__dict__[key] = data_value
                query_parameter_l.__dict__[key] = data_value
                    
            query_parameter.save()
            query_parameter_l.save()
            serializers = JobdescriptionSerializer(query_parameter)
            serializers = CandidateSerializer(query_parameter_l)
            return Response(serializers.data,status=  status.HTTP_201_CREATED)
        except:
            return Response(serializers.errors,status = status.HTTP_400_BAD_REQUEST)

class SelectedAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self,request):
        serializers = limit_off(Selected,request, SelectedSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    
    def post(self,request):
        serializers = SelectedSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status= 201)
        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)
    
class SelectedCandidateAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self,request):        
        query= Jobdescription.objects.filter(status__status = "selected")
        list_name = []
        for name in query:
            list_name.append({int(name.id): name.candidate.Candidatename})
        return Response({"selected_candidates":list_name}, status=status.HTTP_200_OK)

class FollowupAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self,request):
        serializers = limit_off(Followup,request, FollowupSerializer)
        return Response(serializers, status=status.HTTP_200_OK)
    def post(self,request):
        serializers = FollowupSerializer(data = request.data)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data,status= 201)
        return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)


class exportUsersCsv(APIView):
    def get(self, request):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ['Company Name', 'Representative Name','Email']
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = header
            emails = request.query_params.get("email").split(",") 
            company = request.query_params.get("company").split(",")
            Representative = request.query_params.get("representative").split(",")
            init=0
            for row_num, email in enumerate(emails,2):
                sheet[f"A{row_num}"] = company[init]
                sheet[f"B{row_num}"] = Representative[init]
                sheet[f"C{row_num}"] = email
                init+=1
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="BitSquadEmployeeAttendanceData.xlsx"'
            response.write(buffer.getvalue())
            return response
        except:
            return Response({"status": "An error ocured. Try again!!!"}, status=status.HTTP_400_BAD_REQUEST)
